import json
import logging
import random
import secrets

from django.conf import settings
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.db import transaction
from django.http import JsonResponse
from django.middleware.csrf import get_token
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from apps.accesscontrol.permissions import require_roles
from apps.accesscontrol.scoped_service import scope_queryset

from apps.academics.models import (
    Institution,
    Section,
    Subject,
    ClassRoom,
    TeachingAssignment,
    District,
)
from .models import (
    StudentRegistrationRecord,
    OTPVerification,
    DeviceSession,
    JoinCode,
)
from .services import send_otp_async

User = get_user_model()
logger = logging.getLogger(__name__)


# =========================================================
# INTERNAL SERVICE: Teacher assignment creation
# =========================================================

def assign_teacher_to_classes(teacher, subject, institution):
    classrooms = ClassRoom.objects.filter(
        institution=institution,
        name__in=["6", "7", "8", "9", "10"],
    )
    created_count = 0
    for classroom in classrooms:
        for section in Section.objects.filter(classroom=classroom):
            _, created = TeachingAssignment.objects.get_or_create(
                teacher=teacher,
                subject=subject,
                section=section,
            )
            if created:
                created_count += 1

    logger.info(
        "Teacher id=%s assigned to %d sections for subject '%s' in '%s'.",
        teacher.id,
        created_count,
        subject.name,
        institution.name,
    )

_assign_teacher_to_classes = assign_teacher_to_classes


def _create_device_session(request, user):
    if not request.session.session_key:
        request.session.save()

    DeviceSession.objects.filter(user=user).delete()
    DeviceSession.objects.create(
        user=user,
        device_fingerprint=request.session.session_key,
    )


# =========================================================
# REGISTER
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def register(request):
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    username = body.get("username", "").strip()
    password = body.get("password", "")
    join_code_value = body.get("join_code", "").strip()

    if not username or not password or not join_code_value:
        return JsonResponse(
            {"error": "username, password and join_code are required"},
            status=400,
        )

    try:
        join_code = JoinCode.objects.select_related(
            "institution", "section", "district", "subject",
        ).get(code=join_code_value)
    except JoinCode.DoesNotExist:
        return JsonResponse({"error": "Invalid join code"}, status=400)

    if not join_code.is_valid():
        return JsonResponse({"error": "Expired or already used join code"}, status=400)

    if User.objects.filter(username=username).exists():
        return JsonResponse({"error": "Username already exists"}, status=400)

    role = join_code.role
    institution = join_code.institution
    section = join_code.section

    district = None
    if institution:
        district = institution.district.name
    elif section and section.classroom and section.classroom.institution:
        district = section.classroom.institution.district.name
    elif join_code.district:
        district = join_code.district.name

    with transaction.atomic():
        user = User.objects.create_user(
            username=username,
            password=password,
            role=role,
            institution=institution,
            section=section,
            district=district,
        )

        if role == "TEACHER" and join_code.subject and institution:
            assign_teacher_to_classes(user, join_code.subject, institution)

        join_code.mark_as_used()

    logger.info(
        "New user registered: id=%s username=%s role=%s",
        user.id, user.username, user.role,
    )

    return JsonResponse({
        "id": user.id,
        "public_id": user.public_id,
        "username": user.username,
        "role": user.role,
        "institution": user.institution.name if user.institution else None,
        "section": user.section.name if user.section else None,
    })


# =========================================================
# STUDENT SELF REGISTRATION
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def student_register(request):
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    code     = body.get("registration_code", "").strip()
    username = body.get("username", "").strip()
    password = body.get("password", "")
    dob      = body.get("dob", "").strip()

    if not all([code, username, password, dob]):
        return JsonResponse({"error": "Missing required fields"}, status=400)

    try:
        record = StudentRegistrationRecord.objects.select_related(
            "section__classroom__institution",
        ).get(registration_code=code)
    except StudentRegistrationRecord.DoesNotExist:
        return JsonResponse({"error": "Invalid registration code"}, status=400)

    if record.is_registered:
        return JsonResponse({"error": "Registration code already used"}, status=400)

    if str(record.dob) != str(dob):
        return JsonResponse({"error": "Date of birth does not match"}, status=400)

    if User.objects.filter(username=username).exists():
        return JsonResponse({"error": "Username already exists"}, status=400)

    with transaction.atomic():
        user = User.objects.create_user(
            username=username,
            password=password,
            role="STUDENT",
            institution=record.section.classroom.institution,
            section=record.section,
        )
        record.is_registered = True
        record.linked_user = user
        record.save(update_fields=["is_registered", "linked_user"])

    return JsonResponse({
        "id": user.id,
        "public_id": user.public_id,
        "username": user.username,
        "section": user.section.name,
        "institution": user.institution.name,
    })


# =========================================================
# LOGIN
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def login_view(request):

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    user = authenticate(
        username=body.get("username", ""),
        password=body.get("password", ""),
    )

    if not user:
        return JsonResponse({"error": "Invalid credentials"}, status=401)

    if user.role in ["STUDENT", "ADMIN"]:
        login(request, user)
        _create_device_session(request, user)
        return JsonResponse({
            "otp_required": False,
            "id": user.id,
            "username": user.username,
            "role": user.role,
        })

    otp_code = str(secrets.randbelow(900000) + 100000)  # cryptographically secure 6-digit OTP
    OTPVerification.objects.filter(user=user).delete()
    OTPVerification.objects.create(user=user, otp_code=otp_code)

    # Fire-and-forget OTP delivery — runs in background thread.
    # Login response returns instantly; OTP arrives within seconds.
    channel = send_otp_async(user, otp_code)

    response_data = {
        "otp_required": True,
        "otp_channel":  channel,   # "sms", "email", or "log"
        "id":           user.id,
        "username":     user.username,
        "role":         user.role,
    }

    # Only expose otp_code in response when DEBUG=True (dev only)
    if settings.DEBUG:
        response_data["otp_code"] = otp_code

    return JsonResponse(response_data)


# =========================================================
# VERIFY OTP
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def verify_otp(request):
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    username  = body.get("username", "").strip()
    otp_input = body.get("otp", "").strip()

    if not username or not otp_input:
        return JsonResponse({"error": "Missing username or OTP"}, status=400)

    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return JsonResponse({"error": "Invalid credentials"}, status=400)

    otp_record = (
        OTPVerification.objects
        .filter(user=user)
        .order_by("-created_at")
        .first()
    )

    if not otp_record or otp_record.is_expired():
        return JsonResponse({"error": "OTP expired or not found"}, status=400)

    if otp_record.attempt_count >= 5:
        return JsonResponse({"error": "Too many attempts. Request a new OTP."}, status=429)

    if otp_record.otp_code != otp_input:
        otp_record.attempt_count += 1
        otp_record.last_attempt_at = timezone.now()
        otp_record.save(update_fields=["attempt_count", "last_attempt_at"])
        return JsonResponse({"error": "Invalid OTP"}, status=400)

    login(request, user)
    _create_device_session(request, user)

    otp_record.is_verified = True
    otp_record.save(update_fields=["is_verified"])

    return JsonResponse({"success": True, "role": user.role})


# =========================================================
# RESEND OTP
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def resend_otp(request):
    """
    Resend OTP for a user who has already attempted login.
    Generates a new 6-digit code, deletes old ones, and re-sends.

    Rate limit: Only allowed once every 60 seconds per user (enforced
    by checking the most recent OTPVerification created_at).
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    username = body.get("username", "").strip()
    if not username:
        return JsonResponse({"error": "username is required"}, status=400)

    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return JsonResponse({"error": "Invalid credentials"}, status=400)

    # Rate limit: 60s cooldown
    recent = (
        OTPVerification.objects
        .filter(user=user)
        .order_by("-created_at")
        .first()
    )
    if recent:
        elapsed = (timezone.now() - recent.created_at).total_seconds()
        if elapsed < 60:
            wait = int(60 - elapsed)
            return JsonResponse(
                {"error": f"Please wait {wait} seconds before requesting a new OTP.", "retry_after": wait},
                status=429,
            )

    # Generate new OTP
    otp_code = str(secrets.randbelow(900000) + 100000)
    OTPVerification.objects.filter(user=user).delete()
    OTPVerification.objects.create(user=user, otp_code=otp_code)

    channel = send_otp_async(user, otp_code)

    response_data = {
        "resent":      True,
        "otp_channel": channel,
    }
    if settings.DEBUG:
        response_data["otp_code"] = otp_code

    logger.info("OTP resent for user %s via %s", username, channel)
    return JsonResponse(response_data)


# =========================================================
# LOGOUT
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def logout_view(request):
    if request.user.is_authenticated:
        DeviceSession.objects.filter(user=request.user).delete()
        logger.info("User id=%s logged out.", request.user.id)
    logout(request)
    return JsonResponse({"success": True})


# =========================================================
# ME
# =========================================================

@require_http_methods(["GET"])
def me(request):
    if not request.user.is_authenticated:
        return JsonResponse({"authenticated": False})

    user = (
        User.objects
        .select_related("institution", "section", "section__classroom")
        .get(id=request.user.id)
    )

    institution_name = user.institution.name if user.institution else None
    district_name = user.district if user.district else None
    
    if institution_name and district_name and institution_name.endswith(district_name):
        institution_name = institution_name[:-len(district_name)].strip()

    return JsonResponse({
        "authenticated":    True,
        "id":               user.id,
        "public_id":        user.public_id,
        "username":         user.username,
        "role":             user.role,
        "first_name":       user.first_name,
        "middle_name":      user.middle_name,
        "last_name":        user.last_name,
        "display_name":     user.display_name,
        "email":            user.email,
        "mobile_primary":   user.mobile_primary,
        "mobile_secondary": user.mobile_secondary,
        "profile_complete": user.profile_complete,
        "institution":      institution_name,
        "institution_id":   user.institution.id  if user.institution else None,
        "section":          f"Class {user.section.classroom.name}-{user.section.name}" if user.section and hasattr(user.section, 'classroom') else (user.section.name if user.section else None),
        "section_id":       user.section.id      if user.section     else None,
        "district":         district_name,
    })


# =========================================================
# CSRF TOKEN
# =========================================================

@require_http_methods(["GET"])
def csrf_token_view(request):
    return JsonResponse({"csrfToken": get_token(request)})


# =========================================================
# VALIDATE JOIN CODE
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def validate_join_code(request):
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    join_code_value = body.get("join_code", "").strip()

    if not join_code_value:
        return JsonResponse({"error": "join_code is required"}, status=400)

    try:
        join_code = JoinCode.objects.select_related(
            "institution", "section", "district", "subject"
        ).get(code=join_code_value)
    except JoinCode.DoesNotExist:
        return JsonResponse({"error": "Invalid join code"}, status=400)

    if not join_code.is_valid():
        return JsonResponse({"error": "Expired or already used join code"}, status=400)

    return JsonResponse({
        "valid": True,
        "role": join_code.role,
        "institution": join_code.institution.name if join_code.institution else None,
        "section": join_code.section.name if join_code.section else None,
        "district": join_code.district.name if join_code.district else None,
        "subject": join_code.subject.name if join_code.subject else None,
    })


# =========================================================
# SCOPED LIST ENDPOINTS
# =========================================================

@require_roles(["ADMIN", "OFFICIAL", "PRINCIPAL", "TEACHER"])
@require_http_methods(["GET"])
def users(request):
    """
    GET /api/v1/accounts/users/

    Returns users scoped to the calling user's role:
      ADMIN     → all users
      OFFICIAL  → users in their district
      PRINCIPAL → users in their institution
      TEACHER   → only STUDENT users in their institution (no other staff)

    FIX 2026-03-18: TEACHER was excluded from this endpoint entirely,
    causing a 403 on /teacher/users. Added TEACHER with a restricted
    scope — they can only see students, not staff.
    """
    user = request.user

    if user.role == "TEACHER":
        # Teachers see only students at their institution — never other staff
        if not user.institution:
            return JsonResponse([], safe=False)
        queryset = User.objects.filter(
            role="STUDENT",
            institution=user.institution,
        ).order_by("username")
        data = list(queryset.values("id", "username", "role", "public_id"))
        return JsonResponse(data, safe=False)

    queryset = scope_queryset(request.user, User.objects.all())
    data = list(queryset.values("id", "username", "role", "public_id", "district"))
    return JsonResponse(data, safe=False)


@require_roles(["ADMIN", "OFFICIAL", "PRINCIPAL"])
@require_http_methods(["GET"])
def institutions_list(request):
    cache_key = f"institutions_list:{request.user.id}"
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse(cached, safe=False)

    queryset = scope_queryset(
        request.user,
        Institution.objects.select_related("district").order_by("name"),
    )
    data = list(queryset.values("id", "name", "district__name"))
    cache.set(cache_key, data, timeout=600)
    return JsonResponse(data, safe=False)


@require_roles(["ADMIN", "OFFICIAL", "PRINCIPAL", "TEACHER"])
@require_http_methods(["GET"])
def sections_list(request):
    """
    GET /api/v1/accounts/sections/

    FIX 2026-03-18: was returning only {id, name, classroom_id} where
    `name` is just "A" or "B" — all sections look identical in the dropdown.
    Now returns `short_label` ("Class 8-A") and `label`
    ("Class 8-A — School Name") so the frontend can show meaningful options.
    """
    cache_key = f"sections_list:{request.user.id}"
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse(cached, safe=False)

    queryset = scope_queryset(
        request.user,
        Section.objects.select_related(
            "classroom",
            "classroom__institution",
        ).order_by("classroom__name", "name"),
    )

    data = []
    for s in queryset:
        classroom = s.classroom
        institution = classroom.institution if classroom else None
        grade = classroom.name if classroom else "?"
        short_label = f"Class {grade}-{s.name}"
        full_label = (
            f"{short_label} — {institution.name}"
            if institution else short_label
        )
        data.append({
            "id":          s.id,
            "name":        s.name,
            "classroom_id": s.classroom_id,
            "grade":       grade,
            "short_label": short_label,
            "label":       full_label,
        })

    cache.set(cache_key, data, timeout=600)
    return JsonResponse(data, safe=False)


@require_roles(["ADMIN", "OFFICIAL", "PRINCIPAL", "TEACHER"])
@require_http_methods(["GET"])
def subjects_list(request):
    cache_key = f"subjects_list:{request.user.id}"
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse(cached, safe=False)

    queryset = scope_queryset(request.user, Subject.objects.all())
    data = list(queryset.values("id", "name"))
    cache.set(cache_key, data, timeout=600)
    return JsonResponse(data, safe=False)


@require_roles(["ADMIN", "OFFICIAL", "PRINCIPAL"])
@require_http_methods(["GET"])
def teachers(request):
    queryset = scope_queryset(
        request.user,
        User.objects.filter(role="TEACHER"),
    )
    return JsonResponse(
        list(queryset.values("id", "username", "public_id")),
        safe=False,
    )


@require_roles(["ADMIN", "PRINCIPAL", "TEACHER", "OFFICIAL"])
@require_http_methods(["GET"])
def join_codes_list(request):
    queryset = JoinCode.objects.select_related(
        "institution", "section", "district", "subject", "created_by"
    ).order_by("-created_at")

    if request.user.role == "OFFICIAL":
        if not request.user.district:
            return JsonResponse({"error": "No district assigned"}, status=400)
        queryset = queryset.filter(
            institution__district__name=request.user.district
        ) | queryset.filter(
            district__name=request.user.district
        )

    elif request.user.role == "PRINCIPAL":
        if not request.user.institution:
            return JsonResponse({"error": "No institution assigned"}, status=400)
        queryset = queryset.filter(institution=request.user.institution)

    elif request.user.role == "TEACHER":
        if not request.user.institution:
            return JsonResponse({"error": "No institution assigned"}, status=400)
        queryset = queryset.filter(
            institution=request.user.institution,
            role="STUDENT",
        )

    data = [
        {
            "id":          jc.id,
            "code":        jc.code,
            "role":        jc.role,
            "institution": jc.institution.name if jc.institution else None,
            "section":     jc.section.name     if jc.section     else None,
            "district":    jc.district.name    if jc.district    else None,
            "subject":     jc.subject.name     if jc.subject     else None,
            "is_used":     jc.is_used,
            "is_valid":    jc.is_valid(),
            "expires_at":  jc.expires_at.isoformat(),
            "created_at":  jc.created_at.isoformat(),
            "created_by":  jc.created_by.username if jc.created_by else None,
        }
        for jc in queryset
    ]

    return JsonResponse(data, safe=False)


@require_roles(["ADMIN", "PRINCIPAL", "TEACHER", "OFFICIAL"])
@require_http_methods(["POST"])
@csrf_exempt
def create_join_code(request):
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    role = body.get("role", "").upper()
    valid_roles = ["STUDENT", "TEACHER", "PRINCIPAL", "OFFICIAL"]

    if role not in valid_roles:
        return JsonResponse(
            {"error": f"role must be one of: {', '.join(valid_roles)}"},
            status=400,
        )

    if request.user.role == "TEACHER" and role != "STUDENT":
        return JsonResponse({"error": "Teachers can only create STUDENT codes."}, status=403)
    if request.user.role == "PRINCIPAL" and role in ["OFFICIAL", "PRINCIPAL"]:
        return JsonResponse({"error": "Principals can only create STUDENT and TEACHER codes."}, status=403)
    if request.user.role == "OFFICIAL" and role != "PRINCIPAL":
        return JsonResponse({"error": "Officials can only create PRINCIPAL codes."}, status=403)

    institution = None
    section = None
    district = None
    subject = None

    if request.user.role in ("TEACHER", "PRINCIPAL"):
        institution = request.user.institution
        if not institution:
            return JsonResponse({"error": "Your account has no institution assigned."}, status=400)
    elif request.user.role == "OFFICIAL":
        institution_id = body.get("institution_id")
        if institution_id:
            institution = get_object_or_404(Institution, id=institution_id)
            if institution.district.name != request.user.district:
                return JsonResponse({"error": "That institution is not in your district."}, status=403)
    else:
        institution_id = body.get("institution_id")
        if institution_id:
            institution = get_object_or_404(Institution, id=institution_id)

    section_id = body.get("section_id")
    if section_id:
        section = get_object_or_404(Section, id=section_id)
        if institution and section.classroom.institution != institution:
            return JsonResponse({"error": "Section does not belong to this institution."}, status=400)

    district_id = body.get("district_id")
    if district_id:
        district = get_object_or_404(District, id=district_id)
    elif request.user.role == "OFFICIAL" and not district_id:
        try:
            district = District.objects.get(name=request.user.district)
        except District.DoesNotExist:
            pass

    subject_id = body.get("subject_id")
    if subject_id:
        subject = get_object_or_404(Subject, id=subject_id)

    expires_days = min(int(body.get("expires_days", 3)), 30)
    expires_at = timezone.now() + timezone.timedelta(days=expires_days)

    try:
        join_code = JoinCode(
            role=role,
            institution=institution,
            section=section,
            district=district,
            subject=subject,
            created_by=request.user,
            expires_at=expires_at,
        )
        join_code.full_clean()
        join_code.save()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    logger.info("Join code created: id=%s role=%s by user id=%s", join_code.id, join_code.role, request.user.id)

    return JsonResponse({
        "id":          join_code.id,
        "code":        join_code.code,
        "role":        join_code.role,
        "institution": join_code.institution.name if join_code.institution else None,
        "section":     join_code.section.name     if join_code.section     else None,
        "district":    join_code.district.name    if join_code.district    else None,
        "subject":     join_code.subject.name     if join_code.subject     else None,
        "expires_at":  join_code.expires_at.isoformat(),
        "is_valid":    join_code.is_valid(),
    }, status=201)


@require_roles(["ADMIN", "PRINCIPAL", "TEACHER"])
@require_http_methods(["POST"])
@csrf_exempt
def revoke_join_code(request, code_id):
    """
    Mark a join code as used/revoked.

    FIX 2026-03-18: added TEACHER to allowed roles. Teachers can revoke
    STUDENT codes for their own institution. Added scope check to prevent
    a teacher revoking codes from other institutions.
    """
    join_code = get_object_or_404(JoinCode, id=code_id)

    if request.user.role == "PRINCIPAL":
        if not request.user.institution or join_code.institution != request.user.institution:
            return JsonResponse({"error": "Forbidden"}, status=403)

    if request.user.role == "TEACHER":
        if not request.user.institution or join_code.institution != request.user.institution:
            return JsonResponse({"error": "Forbidden"}, status=403)
        if join_code.role != "STUDENT":
            return JsonResponse({"error": "Teachers can only revoke STUDENT codes."}, status=403)

    JoinCode.objects.filter(pk=join_code.pk).update(is_used=True)
    logger.info("Join code id=%s revoked by user id=%s", code_id, request.user.id)

    return JsonResponse({"success": True, "code": join_code.code})


# =========================================================
# JOIN CODE EXCEL EXPORT
# =========================================================

@require_roles(["ADMIN", "PRINCIPAL", "TEACHER", "OFFICIAL"])
@require_http_methods(["GET"])
def export_join_codes(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from django.http import HttpResponse

    queryset = JoinCode.objects.select_related(
        "institution", "section", "district", "subject", "created_by"
    ).order_by("-created_at")

    if request.user.role == "PRINCIPAL":
        if not request.user.institution:
            return JsonResponse({"error": "No institution assigned"}, status=400)
        queryset = queryset.filter(institution=request.user.institution)
    elif request.user.role == "TEACHER":
        if not request.user.institution:
            return JsonResponse({"error": "No institution assigned"}, status=400)
        queryset = queryset.filter(institution=request.user.institution, role="STUDENT")
    elif request.user.role == "OFFICIAL":
        if not request.user.district:
            return JsonResponse({"error": "No district assigned"}, status=400)
        queryset = (
            queryset.filter(institution__district__name=request.user.district) |
            queryset.filter(district__name=request.user.district)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Join Codes"

    DARK = "1E293B"; LIGHT1 = "FFFFFF"; LIGHT2 = "F1F5F9"
    BRAND = "3B82F6"; MUTED = "64748B"

    header_fill = PatternFill("solid", start_color=DARK)
    header_font = Font(bold=True, color="F8FAFC", name="Arial", size=11)
    body_font   = Font(name="Arial", size=10, color="1E293B")
    code_font   = Font(name="Courier New", size=10, color=BRAND, bold=True)
    muted_font  = Font(name="Arial", size=9, color=MUTED)

    border_side = Side(style="thin", color="E2E8F0")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    headers    = ["#", "Role", "Join Code", "Institution / District", "Section", "Subject", "Expires", "Status", "Created By"]
    col_widths = [4,    12,     22,          36,                        14,        18,         14,        10,       16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    row_fills = [LIGHT1, LIGHT2]
    for ri, jc in enumerate(queryset, 2):
        fill   = PatternFill("solid", start_color=row_fills[(ri - 2) % 2])
        status = "Active" if jc.is_valid() else ("Used" if jc.is_used else "Expired")
        for_val = jc.institution.name if jc.institution else (jc.district.name if jc.district else "—")
        row_data = [
            ri - 1, jc.role, jc.code, for_val,
            jc.section.name if jc.section else "—",
            jc.subject.name if jc.subject else "—",
            jc.expires_at.strftime("%d %b %Y"), status,
            jc.created_by.username if jc.created_by else "—",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border; cell.fill = fill
            if ci == 3:
                cell.font = code_font; cell.alignment = left
            elif ci in (1, 2, 7, 8):
                cell.font = body_font; cell.alignment = center
            else:
                cell.font = body_font; cell.alignment = left
        ws.row_dimensions[ri].height = 20

    total_rows = queryset.count()
    ws.cell(row=total_rows + 2, column=1,
            value=f"Total: {total_rows} codes | Exported by: {request.user.username}").font = muted_font

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"gyangrit_join_codes_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Access-Control-Expose-Headers"] = "Content-Disposition"
    return response


# =========================================================
# EMAIL A SINGLE JOIN CODE
# =========================================================

@require_roles(["ADMIN", "PRINCIPAL", "TEACHER", "OFFICIAL"])
@require_http_methods(["POST"])
@csrf_exempt
def email_join_code(request, code_id):
    import json as _json
    from django.core.mail import send_mail
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError

    try:
        body = _json.loads(request.body)
    except (_json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    recipient = body.get("email", "").strip()
    if not recipient:
        return JsonResponse({"error": "email is required"}, status=400)

    try:
        validate_email(recipient)
    except ValidationError:
        return JsonResponse({"error": "Invalid email address"}, status=400)

    join_code = get_object_or_404(JoinCode, id=code_id)

    if request.user.role == "PRINCIPAL":
        if join_code.institution != request.user.institution:
            return JsonResponse({"error": "Forbidden"}, status=403)
    elif request.user.role == "TEACHER":
        if join_code.institution != request.user.institution or join_code.role != "STUDENT":
            return JsonResponse({"error": "Forbidden"}, status=403)
    elif request.user.role == "OFFICIAL":
        if join_code.district and join_code.district.name != request.user.district:
            return JsonResponse({"error": "Forbidden"}, status=403)

    if not join_code.is_valid():
        return JsonResponse({"error": "This code is already used or expired."}, status=400)

    for_label = join_code.institution.name if join_code.institution else (join_code.district.name if join_code.district else "GyanGrit")
    subject_line = f"Your GyanGrit Join Code — {join_code.role.capitalize()}"
    body_text = f"""Hello,

You have been invited to join GyanGrit as a {join_code.role.lower()}.

Your join code is:

    {join_code.code}

Use this code on the registration page at: https://gyangrit.com/register

Details:
  Role:        {join_code.role}
  Institution: {for_label}
  Expires:     {join_code.expires_at.strftime('%d %b %Y')}

This code can only be used once.

— GyanGrit Team
"""
    if settings.DEBUG:
        logger.debug("DEV email preview for join code %s → %s", join_code.code, recipient)
        return JsonResponse({"sent": False, "dev_mode": True, "preview": {"to": recipient, "subject": subject_line, "code": join_code.code}})

    try:
        send_mail(
            subject=subject_line,
            message=body_text,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "noreply@gyangrit.com"),
            recipient_list=[recipient],
            fail_silently=False,
        )
        logger.info("Join code email sent: code=%s to=%s by user=%s", join_code.code, recipient, request.user.id)
        return JsonResponse({"sent": True, "to": recipient})
    except Exception as e:
        logger.error("Failed to send join code email: %s", str(e))
        return JsonResponse({"error": "Email delivery failed. Check server email configuration."}, status=500)


# =========================================================
# COMPLETE PROFILE
# =========================================================

@require_http_methods(["PATCH"])
@csrf_exempt
def complete_profile(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Authentication required"}, status=401)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    user = request.user
    errors = {}
    update_fields = ["profile_complete"]

    if "first_name" in body:
        val = body["first_name"].strip()
        if not val:
            errors["first_name"] = "First name is required."
        elif len(val) < 2:
            errors["first_name"] = "First name must be at least 2 characters."
        else:
            user.first_name = val
            update_fields.append("first_name")

    if "middle_name" in body:
        user.middle_name = body["middle_name"].strip()
        update_fields.append("middle_name")

    if "last_name" in body:
        val = body["last_name"].strip()
        if not val:
            errors["last_name"] = "Last name is required."
        elif len(val) < 2:
            errors["last_name"] = "Last name must be at least 2 characters."
        else:
            user.last_name = val
            update_fields.append("last_name")

    if "email" in body:
        email = body["email"].strip().lower()
        if not email:
            errors["email"] = "Email is required."
        else:
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError as DjangoValidationError
            try:
                validate_email(email)
                user.email = email
                update_fields.append("email")
            except DjangoValidationError:
                errors["email"] = "Enter a valid email address."

    if "mobile_primary" in body:
        raw = body["mobile_primary"].strip()
        digits = "".join(c for c in raw if c.isdigit())
        if not raw:
            errors["mobile_primary"] = "Primary mobile number is required."
        elif len(digits) < 10:
            errors["mobile_primary"] = "Enter a valid 10-digit mobile number."
        else:
            user.mobile_primary = raw
            update_fields.append("mobile_primary")

    if "mobile_secondary" in body:
        raw = body["mobile_secondary"].strip()
        if raw:
            digits = "".join(c for c in raw if c.isdigit())
            if len(digits) < 10:
                errors["mobile_secondary"] = "Enter a valid 10-digit mobile number."
            else:
                user.mobile_secondary = raw
                update_fields.append("mobile_secondary")
        else:
            user.mobile_secondary = ""
            update_fields.append("mobile_secondary")

    if errors:
        return JsonResponse({"errors": errors}, status=400)

    was_complete = user.profile_complete

    is_complete = bool(
        user.first_name.strip()
        and user.last_name.strip()
        and user.mobile_primary.strip()
        and user.email.strip()
    )

    user.profile_complete = is_complete
    user.save(update_fields=update_fields)

    # Trigger Welcome Notification on first completion
    if is_complete and not was_complete:
        from apps.notifications.models import Notification, NotificationType
        
        sch_str = user.institution.name if hasattr(user, "institution") and user.institution else "GyanGrit"
        
        sec_str = ""
        if hasattr(user, "section") and user.section:
            if hasattr(user.section, "classroom") and user.section.classroom:
                sec_str = f"Class {user.section.classroom.name}-{user.section.name}"
            else:
                sec_str = user.section.name

        msg = f"Congratulations {user.first_name} {user.last_name}, you successfully completed your registration as {str(user.role).capitalize()}"
        if sec_str:
            msg += f" in {sec_str}, {sch_str}!"
        else:
            msg += f" at {sch_str}!"

        Notification.send(
            user=user,
            subject="Welcome to GyanGrit!",
            message=msg,
            notification_type=NotificationType.SUCCESS
        )

    logger.info("Profile updated: user=%s complete=%s", user.id, user.profile_complete)

    return JsonResponse({
        "profile_complete":  user.profile_complete,
        "first_name":        user.first_name,
        "middle_name":       user.middle_name,
        "last_name":         user.last_name,
        "display_name":      user.display_name,
        "email":             user.email,
        "mobile_primary":    user.mobile_primary,
        "mobile_secondary":  user.mobile_secondary,
    })


# =========================================================
# BULK JOIN CODE GENERATION
# =========================================================

@require_roles(["ADMIN", "PRINCIPAL", "TEACHER", "OFFICIAL"])
@require_http_methods(["POST"])
@csrf_exempt
def bulk_create_join_codes(request):
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    role = body.get("role", "").upper()
    valid_roles = ["STUDENT", "TEACHER", "PRINCIPAL", "OFFICIAL"]

    if role not in valid_roles:
        return JsonResponse({"error": f"role must be one of: {', '.join(valid_roles)}"}, status=400)

    try:
        count = int(body.get("count", 1))
    except (TypeError, ValueError):
        return JsonResponse({"error": "count must be an integer."}, status=400)

    if count < 1:
        return JsonResponse({"error": "count must be at least 1."}, status=400)
    if count > 100:
        return JsonResponse({"error": "Maximum 100 codes per batch."}, status=400)

    if request.user.role == "TEACHER" and role != "STUDENT":
        return JsonResponse({"error": "Teachers can only create STUDENT codes."}, status=403)
    if request.user.role == "PRINCIPAL" and role in ["OFFICIAL", "PRINCIPAL"]:
        return JsonResponse({"error": "Principals can only create STUDENT and TEACHER codes."}, status=403)
    if request.user.role == "OFFICIAL" and role != "PRINCIPAL":
        return JsonResponse({"error": "Officials can only create PRINCIPAL codes."}, status=403)

    institution = None; section = None; district = None; subject = None

    if request.user.role in ("TEACHER", "PRINCIPAL"):
        institution = request.user.institution
        if not institution:
            return JsonResponse({"error": "Your account has no institution assigned."}, status=400)
    elif request.user.role == "OFFICIAL":
        institution_id = body.get("institution_id")
        if institution_id:
            institution = get_object_or_404(Institution, id=institution_id)
            if institution.district.name != request.user.district:
                return JsonResponse({"error": "That institution is not in your district."}, status=403)
    else:
        institution_id = body.get("institution_id")
        if institution_id:
            institution = get_object_or_404(Institution, id=institution_id)

    section_id = body.get("section_id")
    if section_id:
        section = get_object_or_404(Section, id=section_id)
        if institution and section.classroom.institution != institution:
            return JsonResponse({"error": "Section does not belong to this institution."}, status=400)

    district_id = body.get("district_id")
    if district_id:
        district = get_object_or_404(District, id=district_id)
    elif request.user.role == "OFFICIAL":
        try:
            district = District.objects.get(name=request.user.district)
        except District.DoesNotExist:
            pass

    subject_id = body.get("subject_id")
    if subject_id:
        subject = get_object_or_404(Subject, id=subject_id)

    expires_days = min(int(body.get("expires_days", 3)), 30)
    expires_at = timezone.now() + timezone.timedelta(days=expires_days)

    created_codes = []
    try:
        with transaction.atomic():
            for _ in range(count):
                jc = JoinCode(
                    role=role, institution=institution, section=section,
                    district=district, subject=subject,
                    created_by=request.user, expires_at=expires_at,
                )
                jc.full_clean()
                jc.save()
                created_codes.append(jc)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    logger.info("Bulk join codes created: count=%d role=%s by user=%s", count, role, request.user.id)

    data = [
        {
            "id": jc.id, "code": jc.code, "role": jc.role,
            "institution": jc.institution.name if jc.institution else None,
            "section":     jc.section.name     if jc.section     else None,
            "district":    jc.district.name    if jc.district    else None,
            "subject":     jc.subject.name     if jc.subject     else None,
            "is_used":    jc.is_used, "is_valid": jc.is_valid(),
            "expires_at": jc.expires_at.isoformat(),
            "created_at": jc.created_at.isoformat(),
            "created_by": jc.created_by.username if jc.created_by else None,
        }
        for jc in created_codes
    ]

    return JsonResponse({"created": len(data), "codes": data}, status=201)


# =========================================================
# SYSTEM STATS  (ADMIN only)
# =========================================================

@require_roles(["ADMIN"])
@require_http_methods(["GET"])
def system_stats(request):
    from django.core.cache import cache as _cache
    from django.db.models import Count as _Count, Q as _Q
    from django.utils import timezone
    from apps.content.models import Course, Lesson, LessonProgress
    from apps.assessments.models import Assessment, AssessmentAttempt
    from apps.notifications.models import Broadcast
    from apps.accounts.models import DeviceSession

    _KEY = "admin:system_stats"
    cached = _cache.get(_KEY)
    if cached is not None:
        return JsonResponse(cached)

    User = get_user_model()
    today = timezone.now().date()

    # 1 query instead of 6 — GROUP BY role
    role_rows = (
        User.objects
        .values("role")
        .annotate(cnt=_Count("id"))
    )
    role_map = {r["role"]: r["cnt"] for r in role_rows}
    total_users = sum(role_map.values())

    payload = {
        "users": {
            "total":      total_users,
            "students":   role_map.get("STUDENT",   0),
            "teachers":   role_map.get("TEACHER",   0),
            "principals": role_map.get("PRINCIPAL", 0),
            "officials":  role_map.get("OFFICIAL",  0),
            "admins":     role_map.get("ADMIN",     0),
        },
        "active_sessions": DeviceSession.objects.count(),
        "content": {
            "courses":               Course.objects.count(),
            "lessons":               Lesson.objects.filter(is_published=True).count(),
            "published_assessments": Assessment.objects.filter(is_published=True).count(),
        },
        "activity": {
            "lessons_completed_today":     LessonProgress.objects.filter(completed=True, last_opened_at__date=today).count(),
            "assessments_submitted_today": AssessmentAttempt.objects.filter(submitted_at__date=today).count(),
            "notifications_sent_today":    Broadcast.objects.filter(sent_at__date=today).count(),
        },
    }

    _cache.set(_KEY, payload, timeout=60)
    return JsonResponse(payload)


# =========================================================
# FORGOT PASSWORD
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def forgot_password(request):
    """
    POST /api/v1/accounts/forgot-password/
    Body: { "username": "..." } OR { "email": "..." }

    Generates a stateless HMAC reset token (Django's default_token_generator)
    and emails a branded reset link to the user's registered email.

    Always returns 200 — even if the user doesn't exist or has no email —
    to prevent user enumeration attacks.
    """
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_encode
    from django.utils.encoding import force_bytes
    from .services import send_password_reset_email_async

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    identifier = (body.get("username") or body.get("email") or "").strip()
    if not identifier:
        return JsonResponse({"error": "username or email is required"}, status=400)

    # Look up user by username OR email. Always return the same response either way.
    user = None
    try:
        if "@" in identifier:
            user = User.objects.filter(email__iexact=identifier).first()
        else:
            user = User.objects.get(username=identifier)
    except User.DoesNotExist:
        pass

    if user and user.email:
        token   = default_token_generator.make_token(user)
        uidb64  = urlsafe_base64_encode(force_bytes(user.pk))
        frontend_url = getattr(settings, "FRONTEND_URL", "https://gyangrit.site")
        reset_url = f"{frontend_url}/reset-password/{uidb64}/{token}"

        send_password_reset_email_async(
            user_email=user.email,
            username=user.display_name or user.username,
            reset_url=reset_url,
        )
        logger.info("Password reset email queued for user id=%s", user.id)
    else:
        # Log quietly — don't reveal anything to the caller
        logger.info("forgot_password: no matching user or no email for identifier=%r", identifier)

    # Always 200 — prevents user enumeration
    return JsonResponse({
        "message": "If an account with that username/email exists, a reset link has been sent."
    })


# =========================================================
# RESET PASSWORD (token-based)
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def reset_password(request):
    """
    POST /api/v1/accounts/reset-password/
    Body: { "uidb64": "...", "token": "...", "new_password": "..." }

    Validates the HMAC token, sets a new password, and invalidates all
    existing Django sessions for the user (single-device policy reset).
    """
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    from django.utils.encoding import force_str
    from django.contrib.sessions.models import Session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    uidb64       = body.get("uidb64", "").strip()
    token        = body.get("token",  "").strip()
    new_password = body.get("new_password", "")

    if not uidb64 or not token or not new_password:
        return JsonResponse({"error": "uidb64, token, and new_password are required"}, status=400)

    if len(new_password) < 8:
        return JsonResponse({"error": "Password must be at least 8 characters"}, status=400)

    # Decode uid
    try:
        uid  = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        return JsonResponse({"error": "Invalid reset link"}, status=400)

    # Validate HMAC token (checks expiry via PASSWORD_RESET_TIMEOUT)
    if not default_token_generator.check_token(user, token):
        return JsonResponse({"error": "Reset link is invalid or has expired"}, status=400)

    # Set new password
    user.set_password(new_password)
    user.save(update_fields=["password"])

    # Invalidate all sessions so old sessions can't be replayed
    DeviceSession.objects.filter(user=user).delete()
    # Also flush Django session store for this user (belt-and-suspenders)
    try:
        for session in Session.objects.all():
            data = session.get_decoded()
            if str(data.get("_auth_user_id")) == str(user.pk):
                session.delete()
    except Exception:
        pass  # Non-fatal if session backend doesn't support this

    logger.info("Password reset successful for user id=%s", user.id)
    return JsonResponse({"success": True, "message": "Password updated. Please log in with your new password."})


# =========================================================
# CHANGE PASSWORD (authenticated)
# =========================================================

@require_http_methods(["POST"])
@csrf_exempt
def change_password(request):
    """
    POST /api/v1/accounts/change-password/
    Body: { "old_password": "...", "new_password": "..." }

    Requires an active session. Verifies old password before replacing it.
    Does NOT invalidate the current session — user stays logged in.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Authentication required"}, status=401)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    old_password = body.get("old_password", "")
    new_password = body.get("new_password", "")

    if not old_password or not new_password:
        return JsonResponse({"error": "old_password and new_password are required"}, status=400)

    if len(new_password) < 8:
        return JsonResponse({"error": "New password must be at least 8 characters"}, status=400)

    if not request.user.check_password(old_password):
        return JsonResponse({"error": "Current password is incorrect"}, status=400)

    if old_password == new_password:
        return JsonResponse({"error": "New password must differ from the current one"}, status=400)

    request.user.set_password(new_password)
    request.user.save(update_fields=["password"])

    # Keep the current session alive — update session auth hash so Django
    # doesn't log the user out after password change (session still valid).
    from django.contrib.auth import update_session_auth_hash
    update_session_auth_hash(request, request.user)

    logger.info("Password changed successfully for user id=%s", request.user.id)
    return JsonResponse({"success": True, "message": "Password updated successfully."})

